import os
import re
import copy
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


HEADER_ROW = 17
FIRST_DATA_ROW = 18

DEFAULT_WORKER_TITLE = "근로조건에 관한 확인서(강사용)"
DEFAULT_CHIEF_TITLE = "근로조건에 관한 확인서(주임용)"
DEFAULT_HR_EMAIL = "kli_hr@konkuk.ac.kr"

APP_TITLE = "글로싸인 계약서 대량전송 생성기"
APP_SUBTITLE = "건국대학교 언어교육원 · 계약서 대량전송 엑셀 생성 도구"

COLOR_BG = "#F5F7F4"
COLOR_SURFACE = "#FFFFFF"
COLOR_SURFACE_ALT = "#EEF3EF"
COLOR_PRIMARY = "#0F5D3F"
COLOR_PRIMARY_DEEP = "#0B4A32"
COLOR_ACCENT = "#C9A227"
COLOR_TEXT = "#1E2A22"
COLOR_MUTED = "#5C6B63"
COLOR_BORDER = "#D7E0DA"
COLOR_SUCCESS = "#1E7A4A"
COLOR_DANGER = "#B23A3A"

TEXT_FIELD_ALIASES = {
    "semester": ["해당학기"],
    "teacher_name": ["강사명(A)", "강사명"],
    "start_date": ["계약시작날짜"],
    "end_date": ["계약종료날짜"],
    "fh": ["시수+간주시간(FH)", "시수+간주시간"],
    "f": ["시수(F)", "시수"],
    "h": ["간주시간(H)", "간주시간"],
    "y": ["확정 기본급(수정 1의자리 올림)(Y)", "확정된 기본급(수정 1의자리 올림)", "확정 기본급"],
    "aj": ["총 표준 근로시간(AJ)", "총표준근로시간(AJ)", "겨울학기 총 표준 근로시간", "총 표준 근로시간"],
    "al": ["총시수(AL)", "총시수"],
    "am": ["보정시수(AM)", "보정시수"],
    "ak": ["총 수업준비 간주시간(AK)", "총수업준비간주시간(AK)", "겨울학기 총 수업준비 간주시간", "총 수업준비 간주시간"],
}

PAYROLL_FIELD_CANDIDATES = {
    "name": ["강사명", "이름", "성명"],
    "f": ["시수"],
    "h": ["간주시간"],
    "y": ["확정된 기본급(수정 1의자리 올림)", "확정 기본급(수정 1의자리 올림)", "확정된 기본급"],
    "aj": ["겨울학기 총 표준 근로시간", "총 표준 근로시간"],
    "ak": ["겨울학기 총 수업준비 간주시간", "총 수업준비 간주시간"],
    "al": ["총시수"],
    "am": ["보정시수"],
    "email": ["이메일", "email", "e-mail"],
    "phone": ["전화번호", "휴대폰번호", "핸드폰번호", "연락처"],
}
ROLE_FIELD_CANDIDATES = ["직책", "보직", "261보직", "254보직", "253보직", "역할", "구분"]

SHEET_NAME_CANDIDATES = {
    "payroll": ["근로계약기초자료", "기초자료", "data"],
    "roster": ["명단", "연락처", "list"],
}


def normalize_header(value: object) -> str:
    text = str(value or "")
    text = text.replace("\n", " ").replace("\r", " ").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def normalize_name(value: object) -> str:
    text = str(value or "").replace("\u00a0", " ")
    text = re.sub(r"\s+", "", text)
    return text.strip().lower()


def normalize_phone(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def to_number(value: object, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if text == "":
        return default
    try:
        return float(text)
    except ValueError:
        return default


def display_number(value: object) -> object:
    if value is None or value == "":
        return ""
    n = to_number(value, 0)
    if math.isclose(n, round(n)):
        return int(round(n))
    return n


def locate_sheet(workbook, candidates: List[str]) -> Optional[Worksheet]:
    normalized = {normalize_header(ws.title): ws for ws in workbook.worksheets}
    for name in candidates:
        ws = normalized.get(normalize_header(name))
        if ws:
            return ws
    return None


def get_sheet_names(excel_path: str) -> List[str]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def find_recommended_sheet_name(sheet_names: List[str], candidates: List[str]) -> Optional[str]:
    normalized_map = {normalize_header(name): name for name in sheet_names}
    for candidate in candidates:
        found = normalized_map.get(normalize_header(candidate))
        if found:
            return found
    for original in sheet_names:
        title = normalize_header(original)
        for candidate in candidates:
            if normalize_header(candidate) in title:
                return original
    return sheet_names[0] if sheet_names else None


def build_header_map(ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
    result = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(header_row, col).value)
        if header:
            result[header] = col
    return result


def find_column(header_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
    normalized_candidates = [normalize_header(c) for c in candidates]
    for candidate in normalized_candidates:
        if candidate in header_map:
            return header_map[candidate]
    for header, col in header_map.items():
        for candidate in normalized_candidates:
            if candidate and candidate in header:
                return col
    return None


def find_contact_header_row(ws: Worksheet) -> int:
    best_row = 1
    best_score = -1
    for row in range(1, min(ws.max_row, 5) + 1):
        header_map = build_header_map(ws, header_row=row)
        score = 0
        if find_column(header_map, ["이름", "성명", "강사명"]):
            score += 1
        if find_column(header_map, ["이메일", "email"]):
            score += 1
        if find_column(header_map, ["전화번호", "휴대폰번호", "핸드폰번호"]):
            score += 1
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def read_contact_book(contact_path: str) -> Dict[str, Dict[str, str]]:
    if not contact_path or not os.path.exists(contact_path):
        return {}

    wb = load_workbook(contact_path, data_only=True)
    best_ws = None
    best_row = 1
    best_score = -1

    for ws in wb.worksheets:
        row = find_contact_header_row(ws)
        header_map = build_header_map(ws, row)
        score = 0
        if find_column(header_map, ["이름", "성명", "강사명"]):
            score += 1
        if find_column(header_map, ["이메일", "email"]):
            score += 1
        if find_column(header_map, ["전화번호", "휴대폰번호", "핸드폰번호"]):
            score += 1
        if score > best_score:
            best_score = score
            best_ws = ws
            best_row = row

    if best_ws is None or best_score <= 0:
        return {}

    header_map = build_header_map(best_ws, best_row)
    name_col = find_column(header_map, ["이름", "성명", "강사명"])
    email_col = find_column(header_map, ["이메일", "email"])
    phone_col = find_column(header_map, ["전화번호", "휴대폰번호", "핸드폰번호"])
    if not name_col:
        return {}

    result = {}
    for r in range(best_row + 1, best_ws.max_row + 1):
        raw_name = best_ws.cell(r, name_col).value
        norm_name = normalize_name(raw_name)
        if not norm_name:
            continue
        result[norm_name] = {
            "name": str(raw_name or "").strip(),
            "email": str(best_ws.cell(r, email_col).value or "").strip() if email_col else "",
            "phone": normalize_phone(best_ws.cell(r, phone_col).value) if phone_col else "",
        }
    return result


def read_roster_sheet_contact_map(workbook) -> Dict[str, Dict[str, str]]:
    roster_ws = locate_sheet(workbook, SHEET_NAME_CANDIDATES["roster"])
    if roster_ws is None:
        return {}

    header_row = find_contact_header_row(roster_ws)
    header_map = build_header_map(roster_ws, header_row)
    name_col = find_column(header_map, ["이름", "성명", "강사명"])
    email_col = find_column(header_map, ["이메일", "email"])
    phone_col = find_column(header_map, ["전화번호", "휴대폰번호", "핸드폰번호"])
    if not name_col:
        return {}

    result = {}
    for r in range(header_row + 1, roster_ws.max_row + 1):
        raw_name = roster_ws.cell(r, name_col).value
        norm_name = normalize_name(raw_name)
        if not norm_name:
            continue
        result[norm_name] = {
            "name": str(raw_name or "").strip(),
            "email": str(roster_ws.cell(r, email_col).value or "").strip() if email_col else "",
            "phone": normalize_phone(roster_ws.cell(r, phone_col).value) if phone_col else "",
        }
    return result


def read_payroll_rows(
    payroll_path: str,
    selected_sheet_name: str,
    external_contact_map: Dict[str, Dict[str, str]],
) -> Tuple[List[Dict[str, object]], List[str], str]:
    wb = load_workbook(payroll_path, data_only=True)

    if not selected_sheet_name:
        raise ValueError("급여 시트가 선택되지 않았습니다. 급여엑셀을 다시 선택하거나 시트를 다시 선택하세요.")
    if selected_sheet_name not in wb.sheetnames:
        raise ValueError(f"선택한 급여 시트 '{selected_sheet_name}'를 파일에서 찾을 수 없습니다.")

    payroll_ws = wb[selected_sheet_name]
    roster_contact_map = read_roster_sheet_contact_map(wb)

    header_map = build_header_map(payroll_ws, 1)
    col = {key: find_column(header_map, candidates) for key, candidates in PAYROLL_FIELD_CANDIDATES.items()}
    role_col = find_column(header_map, ROLE_FIELD_CANDIDATES)

    required = ["name", "f", "h", "y", "aj", "ak", "al", "am"]
    missing = [key for key in required if not col.get(key)]
    if missing:
        readable = ", ".join(missing)
        raise ValueError(f"선택한 급여 시트 '{selected_sheet_name}'에서 필수 열을 찾지 못했습니다: {readable}")

    rows: List[Dict[str, object]] = []
    warnings: List[str] = []

    for r in range(2, payroll_ws.max_row + 1):
        raw_name = payroll_ws.cell(r, col["name"]).value
        name = str(raw_name or "").strip()
        norm_name = normalize_name(raw_name)
        if not norm_name:
            continue

        email = str(payroll_ws.cell(r, col["email"]).value or "").strip() if col.get("email") else ""
        phone = normalize_phone(payroll_ws.cell(r, col["phone"]).value) if col.get("phone") else ""

        if not email or not phone:
            contact_info = external_contact_map.get(norm_name) or roster_contact_map.get(norm_name) or {}
            email = email or contact_info.get("email", "")
            phone = phone or contact_info.get("phone", "")

        role_text = normalize_header(payroll_ws.cell(r, role_col).value) if role_col else ""
        is_chief = "주임" in role_text if role_text else False

        row_data = {
            "source_row": r,
            "name": name,
            "norm_name": norm_name,
            "email": email,
            "phone": phone,
            "f": display_number(payroll_ws.cell(r, col["f"]).value),
            "h": display_number(payroll_ws.cell(r, col["h"]).value),
            "y": display_number(payroll_ws.cell(r, col["y"]).value),
            "aj": display_number(payroll_ws.cell(r, col["aj"]).value),
            "ak": display_number(payroll_ws.cell(r, col["ak"]).value),
            "al": display_number(payroll_ws.cell(r, col["al"]).value),
            "am": display_number(payroll_ws.cell(r, col["am"]).value),
            "is_chief": is_chief,
            "role_text": role_text,
        }
        row_data["fh"] = display_number(to_number(row_data["f"]) + to_number(row_data["h"]))

        if not email:
            warnings.append(f"{name}: 이메일 없음")
        if not phone:
            warnings.append(f"{name}: 휴대폰번호 없음")

        rows.append(row_data)

    return rows, warnings, payroll_ws.title


def copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    for c in range(1, ws.max_column + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)

        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.number_format:
            dst.number_format = src.number_format

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def prepare_template(ws: Worksheet, contract_count: int) -> None:
    needed_rows = contract_count * 2
    current_rows = max(0, ws.max_row - FIRST_DATA_ROW + 1)

    worker_style_row = FIRST_DATA_ROW
    hr_style_row = FIRST_DATA_ROW + 1

    if current_rows < needed_rows:
        ws.insert_rows(ws.max_row + 1, needed_rows - current_rows)
    elif current_rows > needed_rows:
        ws.delete_rows(FIRST_DATA_ROW + needed_rows, current_rows - needed_rows)

    for idx in range(needed_rows):
        target_row = FIRST_DATA_ROW + idx
        source_row = worker_style_row if idx % 2 == 0 else hr_style_row
        copy_row_style(ws, source_row, target_row)
        for c in range(2, max(ws.max_column, 25) + 1):
            ws.cell(target_row, c).value = None

    for row in range(1, min(12, ws.max_row) + 1):
        label = normalize_header(ws[f"C{row}"].value)
        if label == normalize_header("총 계약수"):
            ws[f"D{row}"] = contract_count
        elif label == normalize_header("총 참여자 수"):
            ws[f"D{row}"] = contract_count * 2
        elif label == normalize_header("역할"):
            ws[f"D{row}"] = 2


def detect_template_columns(ws: Worksheet) -> Dict[str, str]:
    result: Dict[str, str] = {}

    for c in range(1, ws.max_column + 1):
        value = ws.cell(HEADER_ROW, c).value
        header = normalize_header(value)
        if not header:
            continue

        letter = ws.cell(HEADER_ROW, c).column_letter

        if "계약" in header and "번호" in header:
            result.setdefault("contract_no", letter)
        elif header == normalize_header("역할"):
            result.setdefault("role", letter)
        elif header == normalize_header("이름") and c <= 9:
            result.setdefault("name", letter)
        elif header == normalize_header("이메일"):
            result.setdefault("email", letter)
        elif "발송언어" in header:
            result.setdefault("language", letter)
        elif "국가" in header and "번호" in header:
            result.setdefault("country_code", letter)
        elif "휴대폰" in header and c <= 9:
            result.setdefault("phone", letter)
        elif "개별 메시지" in header:
            result.setdefault("message", letter)
        elif header == normalize_header("이름") and c in (10, 11):
            result.setdefault("pass_name", letter)
        elif "휴대폰" in header and c in (10, 11):
            result.setdefault("pass_phone", letter)
        elif "기업명" in header:
            result.setdefault("company_name", letter)
        elif "사업자등록번호" in header:
            result.setdefault("business_no", letter)

        for key, aliases in TEXT_FIELD_ALIASES.items():
            if any(normalize_header(alias) == header for alias in aliases):
                result[key] = letter

    return result


def set_if_present(ws: Worksheet, colmap: Dict[str, str], key: str, row: int, value: object) -> None:
    letter = colmap.get(key)
    if letter:
        ws[f"{letter}{row}"] = value


def fill_template(
    template_path: str,
    output_path: str,
    rows: List[Dict[str, object]],
    semester: str,
    start_date: str,
    end_date: str,
    hr_name: str,
    hr_email: str,
    hr_phone: str,
    send_language: str,
    country_code: str,
    contract_title: str = "",
) -> List[str]:
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    prepare_template(ws, len(rows))
    colmap = detect_template_columns(ws)

    if contract_title:
        for r in range(1, min(10, ws.max_row) + 1):
            if normalize_header(ws[f"C{r}"].value) == normalize_header("계약명"):
                ws[f"D{r}"] = contract_title
                break

    missing_optional = []
    for field in ["semester", "teacher_name", "start_date", "end_date", "fh", "y", "al", "aj", "am"]:
        if field not in colmap:
            missing_optional.append(field)

    for idx, item in enumerate(rows, start=1):
        worker_row = FIRST_DATA_ROW + (idx - 1) * 2
        hr_row = worker_row + 1

        set_if_present(ws, colmap, "contract_no", worker_row, idx)
        set_if_present(ws, colmap, "role", worker_row, "근로자")
        set_if_present(ws, colmap, "name", worker_row, item["name"])
        set_if_present(ws, colmap, "email", worker_row, item["email"])
        set_if_present(ws, colmap, "language", worker_row, send_language)
        set_if_present(ws, colmap, "country_code", worker_row, country_code)
        set_if_present(ws, colmap, "phone", worker_row, item["phone"])
        set_if_present(ws, colmap, "semester", worker_row, semester)
        set_if_present(ws, colmap, "teacher_name", worker_row, item["name"])
        set_if_present(ws, colmap, "start_date", worker_row, start_date)
        set_if_present(ws, colmap, "end_date", worker_row, end_date)
        set_if_present(ws, colmap, "fh", worker_row, item["fh"])
        set_if_present(ws, colmap, "f", worker_row, item["f"])
        set_if_present(ws, colmap, "h", worker_row, item["h"])
        set_if_present(ws, colmap, "y", worker_row, item["y"])
        set_if_present(ws, colmap, "aj", worker_row, item["aj"])
        set_if_present(ws, colmap, "al", worker_row, item["al"])
        set_if_present(ws, colmap, "am", worker_row, item["am"])
        set_if_present(ws, colmap, "ak", worker_row, item["ak"])

        set_if_present(ws, colmap, "contract_no", hr_row, idx)
        set_if_present(ws, colmap, "role", hr_row, "인사담당자")
        set_if_present(ws, colmap, "name", hr_row, hr_name)
        set_if_present(ws, colmap, "email", hr_row, hr_email)
        set_if_present(ws, colmap, "language", hr_row, send_language)
        set_if_present(ws, colmap, "country_code", hr_row, country_code)
        set_if_present(ws, colmap, "phone", hr_row, hr_phone)

    wb.save(output_path)
    return missing_optional


def split_rows(rows: List[Dict[str, object]], mode: str) -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    if mode == "worker":
        return rows, []
    if mode == "chief":
        return [], rows

    worker_rows = [r for r in rows if not r.get("is_chief")]
    chief_rows = [r for r in rows if r.get("is_chief")]
    return worker_rows, chief_rows


def make_report_text(
    payroll_sheet_name: str,
    output_base_name: str,
    all_rows: List[Dict[str, object]],
    warnings: List[str],
    worker_rows: List[Dict[str, object]],
    chief_rows: List[Dict[str, object]],
    output_paths: List[str],
    notes: List[str],
) -> str:
    lines = [
        f"사용한 급여 시트: {payroll_sheet_name}",
        f"출력 파일명 기준: {output_base_name}",
        f"전체 대상: {len(all_rows)}명",
        f"강사용 대상: {len(worker_rows)}명",
        f"주임용 대상: {len(chief_rows)}명",
        "",
        "[출력 파일]",
    ]
    for p in output_paths:
        lines.append(f"- {p}")
    if warnings:
        lines.extend(["", "[연락처 경고]"] + [f"- {w}" for w in warnings])
    if notes:
        lines.extend(["", "[참고]"] + [f"- {n}" for n in notes])
    return "\n".join(lines)


class SheetSelectDialog(tk.Toplevel):
    def __init__(self, parent: tk.Tk, sheet_names: List[str], recommended: Optional[str] = None):
        super().__init__(parent)
        self.title("급여 시트 선택")
        self.configure(bg=COLOR_BG)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.result: Optional[str] = None
        self.sheet_names = sheet_names
        self.selected_var = tk.StringVar(value=recommended or (sheet_names[0] if sheet_names else ""))

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

        self.update_idletasks()
        self.geometry(self._center_geometry(parent, 420, 360))
        self.listbox.focus_set()

    def _center_geometry(self, parent, width: int, height: int) -> str:
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (width // 2)
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (height // 2)
        return f"{width}x{height}+{max(x, 100)}+{max(y, 100)}"

    def _build_ui(self):
        outer = ttk.Frame(self, style="Card.TFrame", padding=18)
        outer.pack(fill="both", expand=True, padx=16, pady=16)

        ttk.Label(outer, text="급여 시트를 선택하세요", style="DialogTitle.TLabel").pack(anchor="w")
        ttk.Label(
            outer,
            text="선택한 시트를 기준으로 강사 정보와 급여 데이터를 읽습니다.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(4, 14))

        list_frame = ttk.Frame(outer, style="Surface.TFrame")
        list_frame.pack(fill="both", expand=True)

        self.listbox = tk.Listbox(
            list_frame,
            exportselection=False,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=COLOR_BORDER,
            highlightcolor=COLOR_PRIMARY,
            activestyle="none",
            font=("맑은 고딕", 10),
            bg=COLOR_SURFACE,
            fg=COLOR_TEXT,
            selectbackground=COLOR_PRIMARY,
            selectforeground="#FFFFFF",
        )
        self.listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=scrollbar.set)

        for name in self.sheet_names:
            self.listbox.insert("end", name)

        initial = 0
        if self.selected_var.get() in self.sheet_names:
            initial = self.sheet_names.index(self.selected_var.get())
        self.listbox.selection_set(initial)
        self.listbox.see(initial)
        self.listbox.bind("<Double-Button-1>", lambda e: self._on_confirm())

        btns = ttk.Frame(outer, style="Card.TFrame")
        btns.pack(fill="x", pady=(14, 0))
        btns.columnconfigure(0, weight=1)

        ttk.Button(btns, text="취소", style="Secondary.TButton", command=self._on_cancel).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(btns, text="선택 완료", style="Primary.TButton", command=self._on_confirm).grid(row=0, column=2)

    def _on_confirm(self):
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showwarning("시트 선택", "시트를 하나 선택해 주세요.", parent=self)
            return
        self.result = self.sheet_names[selection[0]]
        self.destroy()

    def _on_cancel(self):
        self.result = None
        self.destroy()

    @classmethod
    def ask(cls, parent: tk.Tk, sheet_names: List[str], recommended: Optional[str] = None) -> Optional[str]:
        dialog = cls(parent, sheet_names, recommended)
        parent.wait_window(dialog)
        return dialog.result


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1120x860")
        self.minsize(1020, 760)
        self.configure(bg=COLOR_BG)

        self.payroll_path = tk.StringVar()
        self.contact_path = tk.StringVar()
        self.worker_template_path = tk.StringVar()
        self.chief_template_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.getcwd())

        self.selected_payroll_sheet = tk.StringVar(value="")
        self.sheet_status_text = tk.StringVar(value="선택된 급여 시트: 없음")

        self.mode = tk.StringVar(value="auto")
        self.semester = tk.StringVar(value="겨울")
        self.start_date = tk.StringVar(value="2025-12-03")
        self.end_date = tk.StringVar(value="2026-02-13")
        self.send_language = tk.StringVar(value="")
        self.country_code = tk.StringVar(value="")

        self.hr_name = tk.StringVar()
        self.hr_email = tk.StringVar(value=DEFAULT_HR_EMAIL)
        self.hr_phone = tk.StringVar()

        self.worker_title = tk.StringVar(value=DEFAULT_WORKER_TITLE)
        self.chief_title = tk.StringVar(value=DEFAULT_CHIEF_TITLE)
        self.output_filename = tk.StringVar(value="글로싸인_대량전송")

        self._configure_style()
        self._build_ui()

    def _configure_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self.option_add("*Font", ("맑은 고딕", 10))

        style.configure(".", background=COLOR_BG, foreground=COLOR_TEXT)
        style.configure("Root.TFrame", background=COLOR_BG)
        style.configure("Hero.TFrame", background=COLOR_PRIMARY)
        style.configure("Card.TFrame", background=COLOR_SURFACE)
        style.configure("Surface.TFrame", background=COLOR_SURFACE)
        style.configure("Section.TLabelframe", background=COLOR_SURFACE, bordercolor=COLOR_BORDER, relief="solid", borderwidth=1)
        style.configure("Section.TLabelframe.Label", background=COLOR_SURFACE, foreground=COLOR_PRIMARY_DEEP, font=("맑은 고딕", 11, "bold"))
        style.configure("Title.TLabel", background=COLOR_PRIMARY, foreground="#FFFFFF", font=("맑은 고딕", 18, "bold"))
        style.configure("Subtitle.TLabel", background=COLOR_PRIMARY, foreground="#E8F1EC", font=("맑은 고딕", 10))
        style.configure("Field.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("맑은 고딕", 10))
        style.configure("Muted.TLabel", background=COLOR_SURFACE, foreground=COLOR_MUTED, font=("맑은 고딕", 9))
        style.configure("DialogTitle.TLabel", background=COLOR_SURFACE, foreground=COLOR_PRIMARY_DEEP, font=("맑은 고딕", 12, "bold"))
        style.configure("Status.TLabel", background=COLOR_SURFACE_ALT, foreground=COLOR_PRIMARY_DEEP, font=("맑은 고딕", 10, "bold"))
        style.configure("InfoValue.TLabel", background=COLOR_SURFACE_ALT, foreground=COLOR_TEXT, font=("맑은 고딕", 10))
        style.configure("Primary.TButton", background=COLOR_PRIMARY, foreground="#FFFFFF", borderwidth=0, focusthickness=0, padding=(16, 10), font=("맑은 고딕", 10, "bold"))
        style.map("Primary.TButton", background=[("active", COLOR_PRIMARY_DEEP), ("pressed", COLOR_PRIMARY_DEEP)])
        style.configure("Secondary.TButton", background=COLOR_SURFACE_ALT, foreground=COLOR_PRIMARY_DEEP, borderwidth=1, padding=(14, 9), font=("맑은 고딕", 10))
        style.map("Secondary.TButton", background=[("active", "#E2EAE4"), ("pressed", "#DDE6DF")])
        style.configure("Browse.TButton", background=COLOR_SURFACE_ALT, foreground=COLOR_PRIMARY_DEEP, padding=(10, 8), font=("맑은 고딕", 9))
        style.map("Browse.TButton", background=[("active", "#E2EAE4"), ("pressed", "#DDE6DF")])
        style.configure("TRadiobutton", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("맑은 고딕", 10))
        style.map("TRadiobutton", background=[("active", COLOR_SURFACE)])
        style.configure("TEntry", fieldbackground="#FFFFFF", foreground=COLOR_TEXT, bordercolor=COLOR_BORDER, lightcolor=COLOR_BORDER, darkcolor=COLOR_BORDER, padding=7)
        style.configure("Horizontal.TSeparator", background=COLOR_BORDER)

    def _build_ui(self):
        root = ttk.Frame(self, style="Root.TFrame", padding=18)
        root.pack(fill="both", expand=True)
        root.rowconfigure(2, weight=1)
        root.columnconfigure(0, weight=1)

        hero = ttk.Frame(root, style="Hero.TFrame", padding=(22, 18))
        hero.grid(row=0, column=0, sticky="ew")

        ttk.Label(hero, text=APP_TITLE, style="Title.TLabel").pack(anchor="w")
        ttk.Label(hero, text=APP_SUBTITLE, style="Subtitle.TLabel").pack(anchor="w", pady=(6, 0))

        content = ttk.Frame(root, style="Root.TFrame")
        content.grid(row=1, column=0, sticky="nsew", pady=(16, 0))
        content.columnconfigure(0, weight=3)
        content.columnconfigure(1, weight=2)
        content.rowconfigure(2, weight=1)

        self._build_file_section(content)
        self._build_option_section(content)
        self._build_hr_section(content)
        self._build_report_section(content)

    def _build_file_section(self, parent):
        frame = ttk.LabelFrame(parent, text="파일 선택", style="Section.TLabelframe", padding=14)
        frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=(0, 10))
        frame.columnconfigure(1, weight=1)

        self._add_file_row(frame, 0, "급여엑셀", self.payroll_path, self._browse_payroll_file, required=True)
        self._add_sheet_status_row(frame, 1)
        self._add_file_row(frame, 2, "연락처엑셀(선택)", self.contact_path, lambda: self._browse_file(self.contact_path), required=False)
        self._add_file_row(frame, 3, "강사용 템플릿", self.worker_template_path, lambda: self._browse_file(self.worker_template_path), required=True)
        self._add_file_row(frame, 4, "주임용 템플릿", self.chief_template_path, lambda: self._browse_file(self.chief_template_path), required=True)
        self._add_file_row(frame, 5, "출력 폴더", self.output_dir, lambda: self._browse_dir(self.output_dir), required=True, is_dir=True)

    def _build_option_section(self, parent):
        frame = ttk.LabelFrame(parent, text="생성 옵션", style="Section.TLabelframe", padding=14)
        frame.grid(row=1, column=0, sticky="nsew", padx=(0, 10), pady=(0, 10))
        for i in range(4):
            frame.columnconfigure(i, weight=1)

        ttk.Label(frame, text="생성 방식", style="Field.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Radiobutton(frame, text="강사용 파일만 생성", value="worker", variable=self.mode).grid(row=0, column=1, sticky="w")
        ttk.Radiobutton(frame, text="주임용 파일만 생성", value="chief", variable=self.mode).grid(row=0, column=2, sticky="w")
        ttk.Radiobutton(frame, text="강사용/주임용 파일 모두 생성", value="auto", variable=self.mode).grid(row=0, column=3, sticky="w")

        fields = [
            ("해당학기", self.semester, 1, 0),
            ("계약시작일", self.start_date, 1, 2),
            ("계약종료일", self.end_date, 2, 0),
            ("발송언어(비우면 기본값)", self.send_language, 2, 2),
            ("국가번호(비우면 기본값)", self.country_code, 3, 0),
            ("강사용 계약명", self.worker_title, 3, 2),
            ("주임용 계약명", self.chief_title, 4, 0),
            ("출력 파일명", self.output_filename, 4, 2),
        ]

        for label, var, row, col in fields:
            ttk.Label(frame, text=label, style="Field.TLabel").grid(row=row, column=col, sticky="w", pady=6, padx=(0, 8))
            ttk.Entry(frame, textvariable=var).grid(row=row, column=col + 1, sticky="ew", pady=6)

    def _build_hr_section(self, parent):
        frame = ttk.LabelFrame(parent, text="인사담당자 고정값", style="Section.TLabelframe", padding=14)
        frame.grid(row=0, column=1, rowspan=2, sticky="nsew", pady=(0, 10))
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="이름", style="Field.TLabel").grid(row=0, column=0, sticky="w", pady=6, padx=(0, 8))
        ttk.Entry(frame, textvariable=self.hr_name).grid(row=0, column=1, sticky="ew", pady=6)

        ttk.Label(frame, text="이메일", style="Field.TLabel").grid(row=1, column=0, sticky="w", pady=6, padx=(0, 8))
        ttk.Entry(frame, textvariable=self.hr_email).grid(row=1, column=1, sticky="ew", pady=6)

        ttk.Label(frame, text="휴대폰번호", style="Field.TLabel").grid(row=2, column=0, sticky="w", pady=6, padx=(0, 8))
        ttk.Entry(frame, textvariable=self.hr_phone).grid(row=2, column=1, sticky="ew", pady=6)

        hint = ttk.Frame(frame, style="Card.TFrame", padding=(0, 10, 0, 0))
        hint.grid(row=3, column=0, columnspan=2, sticky="ew")
        ttk.Separator(hint, orient="horizontal").pack(fill="x", pady=(0, 10))
        ttk.Label(
            hint,
            text="이메일은 기본값으로 자동 입력되어 있으며 필요 시 수정할 수 있습니다.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(0, 4))
        ttk.Label(
            hint,
            text="실행 전 이름·이메일·휴대폰번호가 모두 입력되어 있는지 확인해 주세요.",
            style="Muted.TLabel",
        ).pack(anchor="w")

        action_box = ttk.Frame(frame, style="Card.TFrame", padding=(0, 18, 0, 0))
        action_box.grid(row=4, column=0, columnspan=2, sticky="ew")
        ttk.Button(action_box, text="생성 실행", style="Primary.TButton", command=self.run).pack(fill="x")

    def _build_report_section(self, parent):
        outer = ttk.LabelFrame(parent, text="결과 리포트", style="Section.TLabelframe", padding=14)
        outer.grid(row=2, column=0, columnspan=2, sticky="nsew")
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)

        ttk.Label(
            outer,
            text="생성 결과와 안내 메시지가 여기에 표시됩니다. 마우스 휠 또는 드래그로 이동할 수 있습니다.",
            style="Muted.TLabel",
        ).grid(row=0, column=0, sticky="w", pady=(0, 10))

        text_frame = ttk.Frame(outer, style="Surface.TFrame")
        text_frame.grid(row=1, column=0, sticky="nsew")
        text_frame.rowconfigure(0, weight=1)
        text_frame.columnconfigure(0, weight=1)

        self.report = tk.Text(
            text_frame,
            height=18,
            wrap="word",
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=COLOR_BORDER,
            highlightcolor=COLOR_PRIMARY,
            font=("맑은 고딕", 10),
            bg=COLOR_SURFACE,
            fg=COLOR_TEXT,
            insertbackground=COLOR_TEXT,
            padx=12,
            pady=12,
        )
        self.report.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.report.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.report.configure(yscrollcommand=scrollbar.set)

        self.report.bind("<MouseWheel>", self._on_report_mousewheel)
        self.report.bind("<Button-4>", self._on_report_mousewheel_linux)
        self.report.bind("<Button-5>", self._on_report_mousewheel_linux)
        self.report.bind("<ButtonPress-1>", self._on_report_drag_start)
        self.report.bind("<B1-Motion>", self._on_report_drag_move)

        self._set_report_text("준비되었습니다.\n\n급여엑셀을 선택한 뒤 시트를 지정하고, 필요한 템플릿을 선택한 후 생성 실행을 누르세요.")

    def _add_file_row(self, parent, row, label, variable, browse_cmd, required=False, is_dir=False):
        label_text = f"{label} {'*' if required else ''}"
        ttk.Label(parent, text=label_text, style="Field.TLabel").grid(row=row, column=0, sticky="w", pady=6, padx=(0, 8))

        entry = ttk.Entry(parent, textvariable=variable)
        entry.grid(row=row, column=1, sticky="ew", pady=6, padx=(0, 8))

        button_text = "폴더 선택" if is_dir else "찾기"
        ttk.Button(parent, text=button_text, style="Browse.TButton", command=browse_cmd).grid(row=row, column=2, pady=6)

    def _add_sheet_status_row(self, parent, row):
        box = ttk.Frame(parent, style="Card.TFrame")
        box.grid(row=row, column=1, columnspan=2, sticky="ew", pady=(0, 8))
        box.columnconfigure(0, weight=1)

        inner = ttk.Frame(box, style="Card.TFrame", padding=10)
        inner.pack(fill="x")

        status_box = ttk.Frame(inner, style="Card.TFrame")
        status_box.pack(fill="x")
        label_container = ttk.Frame(status_box, style="Card.TFrame")
        label_container.pack(side="left", fill="x", expand=True)

        ttk.Label(label_container, text="시트 상태", style="Muted.TLabel").pack(anchor="w")
        self.sheet_status_label = ttk.Label(label_container, textvariable=self.sheet_status_text, style="Status.TLabel")
        self.sheet_status_label.pack(anchor="w", fill="x", pady=(4, 0))

        ttk.Button(inner, text="시트 다시 선택", style="Secondary.TButton", command=self._reselect_sheet).pack(side="right")

    def _browse_file(self, var: tk.StringVar):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if path:
            var.set(path)

    def _browse_dir(self, var: tk.StringVar):
        path = filedialog.askdirectory()
        if path:
            var.set(path)

    def _browse_payroll_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if not path:
            return

        try:
            sheet_names = get_sheet_names(path)
            if not sheet_names:
                raise ValueError("선택한 급여엑셀에서 시트를 찾지 못했습니다.")

            recommended = find_recommended_sheet_name(sheet_names, SHEET_NAME_CANDIDATES["payroll"])
            selected = SheetSelectDialog.ask(self, sheet_names, recommended)

            if not selected:
                self.payroll_path.set("")
                self.selected_payroll_sheet.set("")
                self._update_sheet_status()
                self._set_report_text("급여엑셀 선택이 취소되었습니다.")
                return

            self.payroll_path.set(path)
            self.selected_payroll_sheet.set(selected)
            self._update_sheet_status()
            self._set_report_text(f"급여엑셀 선택 완료\n- 파일: {path}\n- 선택 시트: {selected}")
        except Exception as e:
            self.payroll_path.set("")
            self.selected_payroll_sheet.set("")
            self._update_sheet_status()
            messagebox.showerror("급여엑셀 선택 오류", str(e), parent=self)

    def _reselect_sheet(self):
        path = self.payroll_path.get().strip()
        if not path:
            messagebox.showwarning("시트 선택", "먼저 급여엑셀 파일을 선택해 주세요.", parent=self)
            return

        try:
            sheet_names = get_sheet_names(path)
            if not sheet_names:
                raise ValueError("선택한 급여엑셀에서 시트 목록을 읽지 못했습니다.")

            recommended = self.selected_payroll_sheet.get().strip() or find_recommended_sheet_name(
                sheet_names, SHEET_NAME_CANDIDATES["payroll"]
            )
            selected = SheetSelectDialog.ask(self, sheet_names, recommended)
            if not selected:
                return

            self.selected_payroll_sheet.set(selected)
            self._update_sheet_status()
            self._append_report_line(f"\n시트를 다시 선택했습니다: {selected}")
        except Exception as e:
            messagebox.showerror("시트 선택 오류", str(e), parent=self)

    def _update_sheet_status(self):
        sheet_name = self.selected_payroll_sheet.get().strip()
        if sheet_name:
            self.sheet_status_text.set(f"선택된 급여 시트: {sheet_name}")
        else:
            self.sheet_status_text.set("선택된 급여 시트: 없음")

    def _collect_missing_required_fields(self) -> List[str]:
        missing = []

        if not self.payroll_path.get().strip():
            missing.append("급여엑셀")
        if not self.selected_payroll_sheet.get().strip():
            missing.append("급여 시트")

        mode = self.mode.get()
        if mode in ("worker", "auto") and not self.worker_template_path.get().strip():
            missing.append("강사용 템플릿")
        if mode in ("chief", "auto") and not self.chief_template_path.get().strip():
            missing.append("주임용 템플릿")
        if not self.output_dir.get().strip():
            missing.append("출력 폴더")
        if not self.output_filename.get().strip():
            missing.append("출력 파일명")
        if not self.hr_name.get().strip():
            missing.append("인사담당자 이름")
        if not self.hr_email.get().strip():
            missing.append("인사담당자 이메일")
        if not self.hr_phone.get().strip():
            missing.append("인사담당자 휴대폰번호")

        return missing

    def _set_report_text(self, text: str):
        self.report.delete("1.0", "end")
        self.report.insert("1.0", text)
        self.report.see("1.0")

    def _append_report_line(self, text: str):
        self.report.insert("end", text)
        self.report.see("end")

    def _on_report_mousewheel(self, event):
        self.report.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _on_report_mousewheel_linux(self, event):
        if event.num == 4:
            self.report.yview_scroll(-3, "units")
        elif event.num == 5:
            self.report.yview_scroll(3, "units")
        return "break"

    def _on_report_drag_start(self, event):
        self.report.scan_mark(event.x, event.y)

    def _on_report_drag_move(self, event):
        self.report.scan_dragto(event.x, event.y, gain=1)

    def run(self):
        try:
            missing = self._collect_missing_required_fields()
            if missing:
                readable = ", ".join(missing)
                raise ValueError(f"필수 입력값이 비어 있습니다: {readable}")

            contact_map = read_contact_book(self.contact_path.get().strip())

            all_rows, warnings, payroll_sheet_name = read_payroll_rows(
                payroll_path=self.payroll_path.get().strip(),
                selected_sheet_name=self.selected_payroll_sheet.get().strip(),
                external_contact_map=contact_map,
            )

            if not all_rows:
                raise ValueError(f"선택한 급여 시트 '{payroll_sheet_name}'에서 처리할 데이터 행을 찾지 못했습니다.")

            worker_rows, chief_rows = split_rows(all_rows, self.mode.get())
            output_paths: List[str] = []
            notes: List[str] = []
            base_name = self.output_filename.get().strip()

            if self.mode.get() == "worker":
                chief_rows = []
            elif self.mode.get() == "chief":
                worker_rows = []

            if worker_rows:
                suffix = "_강사용" if self.mode.get() == "auto" else ""
                output_path = os.path.join(self.output_dir.get().strip(), f"{base_name}{suffix}.xlsx")
                missing_optional = fill_template(
                    template_path=self.worker_template_path.get().strip(),
                    output_path=output_path,
                    rows=worker_rows,
                    semester=self.semester.get().strip(),
                    start_date=self.start_date.get().strip(),
                    end_date=self.end_date.get().strip(),
                    hr_name=self.hr_name.get().strip(),
                    hr_email=self.hr_email.get().strip(),
                    hr_phone=normalize_phone(self.hr_phone.get().strip()),
                    send_language=self.send_language.get().strip(),
                    country_code=self.country_code.get().strip(),
                    contract_title=self.worker_title.get().strip(),
                )
                output_paths.append(output_path)
                if missing_optional:
                    notes.append("강사용 템플릿에 없는 텍스트 필드: " + ", ".join(missing_optional))

            if chief_rows:
                suffix = "_주임용" if self.mode.get() == "auto" else ""
                output_path = os.path.join(self.output_dir.get().strip(), f"{base_name}{suffix}.xlsx")
                missing_optional = fill_template(
                    template_path=self.chief_template_path.get().strip(),
                    output_path=output_path,
                    rows=chief_rows,
                    semester=self.semester.get().strip(),
                    start_date=self.start_date.get().strip(),
                    end_date=self.end_date.get().strip(),
                    hr_name=self.hr_name.get().strip(),
                    hr_email=self.hr_email.get().strip(),
                    hr_phone=normalize_phone(self.hr_phone.get().strip()),
                    send_language=self.send_language.get().strip(),
                    country_code=self.country_code.get().strip(),
                    contract_title=self.chief_title.get().strip(),
                )
                output_paths.append(output_path)
                if missing_optional:
                    notes.append("주임용 템플릿에 없는 텍스트 필드: " + ", ".join(missing_optional))

            if self.mode.get() == "auto" and not chief_rows:
                notes.append("주임으로 분류된 행이 없어 주임용 파일은 생성하지 않았습니다.")
            if self.mode.get() == "auto" and not worker_rows:
                notes.append("일반 강사 행이 없어 강사용 파일은 생성하지 않았습니다.")

            report = make_report_text(
                payroll_sheet_name=payroll_sheet_name,
                output_base_name=base_name,
                all_rows=all_rows,
                warnings=warnings,
                worker_rows=worker_rows,
                chief_rows=chief_rows,
                output_paths=output_paths,
                notes=notes,
            )
            self._set_report_text(report)
            messagebox.showinfo("완료", "파일 생성이 완료되었습니다.", parent=self)

        except Exception as e:
            messagebox.showerror("오류", str(e), parent=self)


if __name__ == "__main__":
    app = App()
    app.mainloop()